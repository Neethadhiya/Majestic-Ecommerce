import re
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from accounts.models import CustomUser
import random
from category.models import Category, Product, ProductImage, Variant, Offer, Banner, Coupon, UsedCoupon
from cart.models import Cart, NewCartItem, Address, Payment, Orders, OrderProduct
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.contrib.auth import authenticate,login,logout
import requests
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.db.models import Sum
from django.contrib.auth.hashers import make_password
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import xlwt
from django.core.mail import send_mail
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from django.http import HttpResponse
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.contrib.auth.views import PasswordResetView
from django.contrib.messages.views import SuccessMessageMixin

class ResetPasswordView(SuccessMessageMixin, PasswordResetView):
    template_name = 'customer/password_reset.html'
    email_template_name = 'customer/password_reset_email.html'
    subject_template_name = 'customer/password_reset_subject.txt'
    success_message = "We've emailed you instructions for setting your password, " \
                      "if an account exists with the email you entered. You should receive them shortly." \
                      " If you don't receive an email, " \
                      "please make sure you've entered the address you registered with, and check your spam folder."
    success_url = reverse_lazy('customer_signin')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def index(request):
    products         =   Product.objects.filter(is_deleted=False)
    categories       =   Category.objects.all()
    banners          =   Banner.objects.all()
    
    if request.method=='POST':
        product     =   request.POST['product']
    if request.user.is_authenticated :
        user                            =   CustomUser.objects.get(email = request.user)
        user_name                       =   user.first_name
        count_item                      =   NewCartItem.objects.filter(user=request.user).aggregate(Sum('quantity'))['quantity__sum'] or 0
        request.session["item_count"]   =   count_item
        context={
                 'products'      :   products,
                 'euser'         :   True,
                 'categories'    :   categories,
                 'user_name'     :   user_name,
                 'cart'          :   True,
                 'banners'       :   banners,               
              }
        return render(request,'store/index.html',context)
    else:
        context={
                 'categories'   :   categories,
                 'products'     :   products,
                 'cart'         :   True,
                 'banners'      :   banners,         
             }
    return render(request,'store/index.html',context)


def customer_signUp(request):
    if request.method == 'POST':
        # Registration form submitted
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        cpassword = request.POST.get('cpassword')

        # Perform validation checks
        if not first_name:
            messages.error(request, 'First name is required', extra_tags='alert alert-danger')
            return redirect('customer_signUp')

        if not email:
            messages.error(request, 'Email is required', extra_tags='alert alert-danger')
            return redirect('customer_signUp')

        if not phone_number:
            messages.error(request, 'Phone number is required', extra_tags='alert alert-danger')
            return redirect('customer_signUp')

        if not password:
            messages.error(request, 'Password is required', extra_tags='alert alert-danger')
            return redirect('customer_signUp')

        if password != cpassword:
            messages.error(request, 'Passwords do not match', extra_tags='alert alert-danger')
            return redirect('customer_signUp')
        else:
            if CustomUser.objects.filter(phone_number = phone_number).exists():
                messages.error(request,"The phone number is already exists. Try another...", extra_tags='signupphone_number')
                return redirect(customer_signUp)
            if CustomUser.objects.filter(email = email).exists():
                messages.error(request,"The email is already exists. Try another...", extra_tags='signupemail')
                return redirect(customer_signUp)

            # Generate OTP code and send to user's phone number
            otp = str(random.randint(1000, 9999))
            status = send_otp_to_phone(phone_number, otp)
            if status == 'Success':
                request.session['phone_number'] = phone_number
                request.session['otp'] = otp
                request.session['new_user_data'] = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'phone_number': phone_number,
                    'password': password,
                }
                messages.success(request,"OTP sent successfully..")   
                request.session['otp_true'] = True
                return redirect('verify_otp')
            else:
                messages.error(request, 'Failed to send OTP code')
                return redirect('customer_signUp')

    # GET request, display registration form
    return render(request, 'customer/userSignUp.html')

def verify_otp(request):
    if request.method == 'POST':
        # Verify OTP code
        entered_otp = request.POST.get('otp_code')
        if not entered_otp:
            messages.error(request, 'Please enter the otp!', extra_tags='alert alert-danger')
            return redirect('customer_signUp')        
        saved_otp = request.session.get('otp')
        if str(entered_otp) == str(saved_otp):
            # OTP verification successful
            new_user_data = request.session.get('new_user_data')
            user = CustomUser.objects.create_user(
                email     = new_user_data['email'],
                password  = new_user_data['password'],
                first_name = new_user_data['first_name'],
                last_name = new_user_data['last_name'],
                phone_number = new_user_data['phone_number'],           
                  )
            request.session['user_id']  =   user.id
            del request.session['otp']
            del request.session['otp_true']
            del request.session['new_user_data']
            messages.success(request, 'Account created successfully!', extra_tags='alert alert-success')
            return redirect('customer_signin')
        else:
            # OTP verification failed
            del request.session['otp_true']
            messages.error(request, 'Invalid OTP code', extra_tags='alert alert-danger')
            return redirect('customer_signUp')

    # GET request, display OTP verification form
    phone_number = request.session.get('phone_number')
    if phone_number:
        return redirect('customer_signUp')
    else:
        messages.error(request, 'Phone number not found', extra_tags='alert alert-danger')
        return redirect('customer_signUp')

    
def send_otp_to_phone(phone_number, otp_code):
    api_key         =   '162c1276-bd93-11ed-81b6-0200cd936042'
    url             =    f'https://2factor.in/API/V1/{api_key}/SMS/+91{phone_number}/{otp_code}'
    response        =    requests.get(url)
    response_json   =   response.json()
    return response_json['Status']

def customer_signin(request):
    categories      =   Category.objects.filter(is_blocked=True)
    context         =   {'categories':categories}
    if request.method=='POST':
        email=request.POST['email']
        password=request.POST['password']
        user = authenticate(email=email,password=password)
        if user is not None:
            if not user.is_blocked:
                request.session['customer']      =   user.id
                request.session['email']         =   user.email
                request.session['customer_name'] =   user.first_name
                login(request, user)
                return redirect('index') 
            else:
                messages.error(request,"Your account has blocked... Please use another account")
                return redirect('customer_signin')
        else:
            messages.error(request,'Invalid email or password')
            return redirect('customer_signin')
    elif not request.user.is_authenticated:
        return render(request,'customer/userSignin.html',context)
    else:
        return redirect('index')

def customer_logout(request):
    if 'first_name' in request.session:
        request.session.flush()
    logout(request)
    return redirect('index')

@login_required(login_url='customer_signin')
def user_profile(request):
    address         =   Address.objects.filter(user=request.user)
    categories      =   Category.objects.all()
    user_details    =   CustomUser.objects.get(id=request.user.id)
    user_name       =   user_details.first_name
    context={'euser':True,'categories':categories,'user_name':user_name,'user_details':user_details,'addresses':address}
    return render(request,'customer/user_profile.html',context)

def user_edit_profile(request,id):
    user_details        =   CustomUser.objects.get(id=id)
    user                =   CustomUser.objects.get(email = request.user)
    user_name           =   user.first_name
    categories          =   Category.objects.filter(is_blocked=True)
    context={'user_details':user_details,'categories':categories,'user_name':user_name,'euser':True}
    return render(request,'customer/edit_user_profile.html',context)

def update_user(request,id):
    user_details=CustomUser.objects.get(id=id)
    if request.method=='POST':
        user_details.first_name       =   request.POST.get('first-name')
        user_details.last_name        =   request.POST.get('last-name')
        user_details.email            =   request.POST.get('email')
        user_details.phone_number     =   request.POST.get('phone-number')
        user_details.save()
        messages.success(request,'User details updated successfully')
        return redirect('user_profile')
    else:
        return render(request,'customer/edit_user_profile.html')   
        
def editAddress(request,id):
    address     =   Address.objects.get(id=id)
    context     =   {'address':address}
    return render(request,'customer/edit_address.html',context)

def update_address(request,id):
    address=Address.objects.get(id=id)
    if request.method=='POST':
        address.firstname       =   request.POST.get('first_name')
        address.lastname        =   request.POST.get('last_name')
        address.housename       =   request.POST.get('house_name_flat')
        address.locality        =   request.POST.get('Locality')
        address.town            =   request.POST['town_city']
        address.state           =   request.POST.get('state_country')
        address.pincode         =   request.POST.get('postcode_zip')
        address.phonenumber     =   request.POST.get('phone_number')
        address.save()
        messages.success(request,'Address updated successfully')
        return redirect('checkout')
    else:
        return redirect('user_profile')
    
def removeAddress(request, id):
    address = Address.objects.get(id=id)
    address.delete()
    messages.success(request, 'Address deleted successfully')
    if 'checkout' in request.META.get('HTTP_REFERER', ''):
        checkout_url = reverse('checkout')
        next_url = request.GET.get('next', '')
        if next_url:
            next_url = f'?next={next_url}'
        return redirect(checkout_url + next_url)
    else:
        return redirect('user_profile')

def add_address(request):
    firstname       =   ""
    lastname        =   ""
    housename       =   ""
    locality        =   ""
    town            =   ""
    state           =   ""
    pincode         =   ""
    phonenumber     =   ""
    categories=Category.objects.all()
    if request.user.is_authenticated :
        user = CustomUser.objects.get(email = request.user)
        user_name= user.first_name
        if request.method   ==  "POST":
            firstname       =   request.POST.get('first_name')
            lastname        =   request.POST.get('last_name')
            housename       =   request.POST.get('house_name_flat')
            locality        =   request.POST.get('Locality')
            town            =   request.POST['town_city']
            state           =   request.POST.get('state_country')
            pincode         =   request.POST.get('postcode_zip')
            phonenumber     =   request.POST.get('phone_number')
            if firstname == "":
                messages.error(request,'First name must not be empty')
                return redirect(add_address)
            elif lastname == "":
                messages.error(request,"Last Name must not be empty")
                return redirect(add_address)
            elif housename == "":
                messages.error(request,"House Name must not be empty")
                return redirect(add_address)
            elif locality == "":
                messages.error(request,"Locality must not be empty")
                return redirect(add_address)
            elif town == "":
                messages.error(request,"Town / City must not be empty")
                return redirect(add_address)
            elif state== "":
                messages.error(request,"State / Country must not be empty")
                return redirect(add_address)
            elif pincode=="":
                messages.error(request,"Postcode must not be empty")
                return redirect(add_address)
            elif phonenumber=="":
                messages.error(request,"Phone Number must not be empty")
                return redirect(add_address)
            else:
                address = Address(firstname     =   firstname,
                                  lastname      =   lastname,
                                  housename     =   housename,
                                  locality      =   locality,
                                  city          =   town,
                                  state         =   state,
                                  pincode       =   pincode,
                                  phonenumber   =   phonenumber,
                                  user          =   request.user)
                address.save()
                messages.success(request,"Address added successfully")
                checkout_url = request.session['return_address']
                del request.session['return_address']
                return redirect(checkout_url)
        else:
            request.session['return_address'] = request.META.get('HTTP_REFERER', '')
            context={
                    'euser':True,
                    'categories'   :    categories,
                    'user_name'    :    user_name,
                    'first_name'   :    firstname,
                    'last_name'    :    lastname, 
                    'house_name_flat':  housename,
                    'Locality'     :    locality, 
                    'town_city'    :    town, 
                    'state_country':    state,
                    'postcode_zip' :    pincode,
                    'phone_number' :    phonenumber,

                }
            return render(request,'customer/add_address.html',context)

@login_required(login_url='customer_signin')       
def user_wallet(request):
    categories      =   Category.objects.all()
    user_details    =   CustomUser.objects.get(id=request.user.id)
    user_name       =   user_details.first_name
    context         =   {
                           'euser'          :   True,
                           'categories'     :   categories,
                           'user_name'      :   user_name,
                           'user_details'   :   user_details
                        }
    return render(request,'customer/user_wallet.html',context)
        
@login_required(login_url='customer_signin')       
def change_password(request):
    categories=Category.objects.all()
    new_user    =   CustomUser.objects.get(id=request.user.id)
    if request.method == 'POST':
        new_password            =   request.POST.get('new-password')
        confirm_password        =   request.POST.get('confirm-password')
        if new_password == "" :
            messages.error(request,"New password cannot be null")
            return redirect('change_password')
        else:
            if new_password == confirm_password:
                new_user.password = make_password(new_password)
                new_user.save()
                user=authenticate(email=new_user.email,password=new_password)
                login(request,user=user)
                messages.success(request,"Successfully reset the password")
                return redirect('user_profile')
            else:
                messages.error(request,"Password doesnot match")
                return redirect('change_password')
    else:
        context={'euser':True,'categories':categories}
        return render(request,'customer/change_password.html',context)
    
@login_required(login_url='customer_signin')          
def user_orders(request,id):
    orders     =    Orders.objects.filter(is_ordered=True,user=request.user).order_by('-updated_at')
    context    =    {'orders'          :   orders,
                     'euser'           :   True,}
    return render(request,'customer/user_orders.html',context)

@login_required(login_url='customer_signin')          
def user_order_details(request,id):
    order  =  Orders.objects.get(id=id)
    order_products = None
    try:
        order_products   =   OrderProduct.objects.filter(order=id).order_by('-updated_at')
    except:
        pass
    context              =   {
                                'order_products'  :   order_products,
                                'euser'           :   True,

                             }
    return render(request,'customer/user_order_details.html',context)

@login_required(login_url='customer_signin')          
def cancel_product(request):
    total                      =   0
    if request.method == 'POST':
        data                   =   json.loads(request.body)
        order_product_id       =   data.get('order_product_id')
        order_product          =   OrderProduct.objects.get(id=order_product_id)
        order_product.status   =   "Cancelled"
        user                   =   CustomUser.objects.get(email=request.user)
        total                  =   order_product.sub_total
        if user is not None:
            user.wallet       +=   total
            user.save()
        order_product.save()
        id = order_product.id
        response_data          =   {'message': 'Order cancelled','id':id}
        return JsonResponse(response_data, status=200)
    else:
        response_data  =   {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400) 
    
@login_required(login_url='customer_signin')          
def return_product(request):
    total                      =   0
    if request.method == 'POST':
        data                   =   json.loads(request.body)
        order_product_id       =   data.get('order_product_id')
        order_product          =   OrderProduct.objects.get(id=order_product_id)
        order_product.status   =   "Returned"
        user                   =   CustomUser.objects.get(email=request.user)
        total                  =   order_product.order.total
        if user is not None:
            user.wallet = (user.wallet or 0) + total
            user.save()
        order_product.variant.stock     +=    order_product.quantity
        order_product.save()
        id = order_product.id
        response_data          =   {'message': 'Order returned','id':id}
        return JsonResponse(response_data, status=200)
    else:
        response_data  =   {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)  
       
@login_required(login_url='customer_signin')          
def download_invoice(request):
    if request.method == 'POST':
        data         =    json.loads(request.body)
        order_id     =   data.get('orderId')
        orders       =    get_object_or_404(Orders, id=order_id)
        # order_products = OrderProduct.objects.filter(order=orders)
        order_products = orders.order_products.all()
        template     =    get_template('store/pdf_invoice.html')
        total        = sum([products.sub_total for products in order_products])
        
        address      =    orders.address
        context      =    {
                    'date'          :   orders.date_added,
                    'invoice_no'    :   orders.orderid,
                    'address'       :   address,
                    'user_name'     :   address.firstname,
                    'order'         :   orders,
                    'total'         :   total,
                    'order_products': order_products
                    # 'order_products'        :   [order_products]
                }
        html = template.render(context)

        def render_to_pdf(html_string):
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
            if not pdf.err:
                return result.getvalue()
            return None

        pdf_file = render_to_pdf(html)
        if pdf_file:
            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"invoice_{orders.id}.pdf"
            content = f"attachment; filename={filename}"
            response['Content-Disposition'] = content
            return response
        else:
            response_data = {'error': 'Unable to generate PDF'}
            return JsonResponse(response_data, status=400)
    else:
        response_data = {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)
    
@login_required(login_url='customer_signin')          
def download_invoice_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        order_id = data.get('orderId')
        orders = get_object_or_404(Orders, id=order_id)
        order_products = orders.order_products.all()
        total = sum([products.sub_total for products in order_products])
        address = orders.address

        # Create a new workbook and set the active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoice"

        # Set column widths and alignment
        column_widths = [32, 32, 32, 32, 32, 32,32]
        alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')     
        for i, column_width in enumerate(column_widths):
            column_letter = get_column_letter(i+1)
            worksheet.column_dimensions[column_letter].width = column_width
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add header row
        header_row = ['Product Name', 'Size', 'Quantity', 'Price', 'Offer Discount', 'Subtotal', 'Order Status']
        for i, header_text in enumerate(header_row):
            cell = worksheet.cell(row=1, column=i+1)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Add data rows
        for i, order_product in enumerate(order_products):
            row_index = i+2
            worksheet.cell(row=row_index, column=1).value = order_product.variant.product.product_name
            worksheet.cell(row=row_index, column=2).value = order_product.variant.variant_name
            worksheet.cell(row=row_index, column=3).value = order_product.quantity
            worksheet.cell(row=row_index, column=4).value = order_product.variant.price
            worksheet.cell(row=row_index, column=5).value = order_product.offer_discount
            worksheet.cell(row=row_index, column=6).value = order_product.sub_total
            worksheet.cell(row=row_index, column=7).value = order_product.status
            for cell in worksheet[row_index]:
                cell.alignment = alignment
        # Add total row
        total_row_index = len(order_products) + 2
        worksheet.cell(row=total_row_index, column=5).value = "Total"
        worksheet.cell(row=total_row_index, column=6).value = total
        worksheet.cell(row=total_row_index, column=5).font = Font(bold=True)
        worksheet.cell(row=total_row_index, column=6).font = Font(bold=True)
        worksheet.cell(row=total_row_index, column=6).alignment = alignment
        worksheet.cell(row=total_row_index, column=5).alignment = alignment
        for col in range(1, 8):
            cell = worksheet.cell(row=total_row_index, column=col)
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Set response headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="invoice_{orders.id}.xlsx"'

        # Save the workbook to the response
        workbook.save(response)
        return response

    else:
        response_data = {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)
























