
import logging
from django.db import IntegrityError
from django.shortcuts import render,redirect,HttpResponseRedirect,HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib import messages,auth
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
import datetime
from django.utils import timezone
from .models import CustomUser
from category.models import Category,Product
from django.db.models import Q
from category.models import Category, Product, ProductImage, Variant, Offer, Banner, Coupon, UsedCoupon
from cart.models import Cart, NewCartItem, Address, Payment, Orders, OrderProduct
import json
from django.http import JsonResponse
from django.contrib.sessions.models import Session
from django.contrib.auth import get_user_model
# from cart.models import Coupon
from django.db.models import Sum
# from cart.models import Orders
from django.views.decorators.http import require_GET
from django.core import serializers
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, HttpResponse, get_object_or_404
from openpyxl.styles import PatternFill, Font, Alignment


def adminLogin(request):
    if request.method=='POST':
        email           =   request.POST['email']
        password        =   request.POST["password"]
        user_obj        =   CustomUser.objects.filter(email=email)
        if not user_obj:
            messages.warning(request,"Account not found")
            return render(request,'admin/adminLogin.html')
        user_obj=authenticate(email=email,password=password)
        if user_obj and user_obj.is_superuser:
            auth.login(request,user_obj)
            return redirect('adminHome')
        else:
            messages.error(request,"Invalid password")
            return redirect('/')
    else:
        if not request.user.is_authenticated:
            return render(request,'admin/adminLogin.html')
        else:
            return redirect('adminHome/')


@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def adminHome(request):
    # get filter parameters from URL, default to current month
    year = int(request.GET.get('year', datetime.date.today().year))
    month = int(request.GET.get('month', datetime.date.today().month))

    start_date = datetime.date(year, month, 1)
    end_date = start_date.replace(day=28) + datetime.timedelta(days=4)
    end_date = end_date - datetime.timedelta(days=end_date.day)

    sales_data = Orders.objects.filter(
        date_added__gte=start_date,
        date_added__lte=end_date,
        is_ordered=True
    ).values('date_added').annotate(total_sales=Sum('total'))

    chart_data = {
        'labels': [d['date_added'].strftime('%d %b %Y') for d in sales_data],
        'values': [d['total_sales'] for d in sales_data],
    }
    sales_data = sales_data.filter(date_added__year=year)
    years = range(datetime.date.today().year, 2000, -1)
    months = [(1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'), (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'), (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')]
    selected_month = next(m[1] for m in months if m[0] == month)
    context = {
        'chart_data': chart_data,
        'year': year,
        'month': month,
        'selected_month':selected_month,
        'years': years,
        'months': months,
    }
    return render(request, 'admin/adminHome.html', context)


@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def userManagement(request):
    users=CustomUser.objects.filter(is_superuser=False).order_by('-updated_at')
    context={
        'users':users,
        }
    return render(request,'admin/adminUserMngmt.html',context)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def blockCustomer(request):
    if request.method == 'POST':
        data        =   json.loads(request.body)
        user_id     =   data.get('userId')
        user=CustomUser.objects.get(id=user_id)
        if user.is_blocked:
            user.is_blocked=False
            user.save()
        else:
            user.is_blocked=True
            user.save()
            sessions = Session.objects.all()
            for session in sessions:
                session_data = session.get_decoded()
                if session_data.get('_auth_user_id') == user_id:
                    session.delete()
        user.save()
        response_data = {'message': 'User blocked successfully','is_blocked':user.is_blocked}
        return JsonResponse(response_data, status=200)
    else:
        response_data = {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def adminLogout(request):
    if 'username' in request.session:
        request.session.flush()
    logout(request)
    return render(request,'admin/adminLogin.html')

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def searchUser(request):
    query=request.POST['query']
    users_list={
        'users':CustomUser.objects.filter(username__icontains=query)
    }
    return render(request,'searchResult.html',users_list)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_orders(request):
    orders     =    Orders.objects.filter(is_ordered=True).order_by('-updated_at')
    context    =    {'orders':orders}
    return render(request,'admin/admin_orders.html',context)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_order_details(request,id):
    order_products = None
    try:
        order_products   =   OrderProduct.objects.filter(order=id).order_by('-updated_at')
    except:
        pass
    context              =   {
                                'order_products'   :   order_products
                            }
    return render(request,'admin/admin_order_details.html',context)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_order_status(request, id):
    if request.method == 'POST':
        data = json.loads(request.body)
        status = data['status']
        order = OrderProduct.objects.get(id=id)
        order.status = status
        order.save()
        id = order.id
        return JsonResponse({'status': status,'id':id})
    
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)    
def add_banner(request):
    if request.method == "POST":
        banner_title    = request.POST.get('banner_title')
        images          =   request.FILES.get('images')
        if banner_title == '':
            messages.error(request,'Banner title must not be empty')
            return redirect('add_banner')
        
        if images == '':
            messages.error(request,'Image field must not be empty')
            return redirect('add_banner')

        elif Banner.objects.filter(title=banner_title, image=images).exists():
            messages.error(request,'Banner  already exists')
            return redirect('add_banner')

        elif Banner.objects.filter(title=banner_title).exists():
            messages.error(request,'Banner already exists')
            return redirect('add_banner')
        else:
            banner = Banner(title=banner_title, image=images)
            banner.save()
            messages.success(request,'Banner added successfully...')
            return redirect('add_banner')
    else:     
        return render(request,'admin/add_banner.html')   
     
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def banner_management(request):
    banners    =   Banner.objects.all()
    context   =   { 'banners'  :  banners  }
    return render(request,'admin/banner_management.html',context) 
 
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def delete_banner(request,id):
    banner   =  Banner.objects.get(id=id)
    banner.delete()
    return redirect('banner_management')

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_add_coupon(request):
    if request.method == "POST":
        coupon_code = request.POST.get('coupon-code')
        discount = request.POST.get('discount')
        if coupon_code == '':
            messages.error(request,'Coupon code must not be empty',extra_tags='coupon')
            return redirect('admin_add_coupon')
        if coupon_code.strip() == '':
            messages.error(request,'Coupon code must not be empty',extra_tags='coupon')
            return redirect('admin_add_coupon')
        if discount == '':
            messages.error(request,'Discount must not be empty',extra_tags='discount')
            return redirect('admin_add_coupon')

        elif Coupon.objects.filter(discount=discount, coupon_code=coupon_code).exists():
            messages.error(request,'Coupon code already exists',extra_tags='coupon')
            return redirect('admin_add_coupon')

        elif Coupon.objects.filter(discount=discount).exists():
            messages.error(request,'Discount already exists',extra_tags='discount')
            return redirect('admin_add_coupon')

        elif Coupon.objects.filter(coupon_code=coupon_code).exists():
            messages.error(request,'Coupon code already exists',extra_tags='coupon')
            return redirect('admin_add_coupon')

        elif int(discount) <=0 or int(discount) >= 100:
            messages.error(request,'Discount must be between 1 and 99',extra_tags='discount')
            return redirect('admin_add_coupon')

        else:
            newcoupon = Coupon(coupon_code=coupon_code, discount=discount)
            newcoupon.save()
            messages.success(request,'Coupon added successfully...')
            return redirect('admin_add_coupon')
    else:
        return render(request,'admin/add_coupon.html')
    
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def display_couponlist(request):
    coupons = Coupon.objects.all().order_by('-id')
    return render(request,'admin/display_coupons.html',{'coupons':coupons}) 
   
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def block_coupon(request):
    if request.method == 'POST':
        data          =   json.loads(request.body)
        coupon_id     =   data.get('couponId')
        coupon        =   Coupon.objects.get(id=coupon_id)
        if coupon.is_active:
            coupon.is_active=False
            coupon.save()
        else:
            coupon.is_active=True
            coupon.save()
        response_data = {'message': 'User blocked successfully','is_active':coupon.is_active}
        return JsonResponse(response_data, status=200)
    else:
        response_data = {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)    
def add_category_offers(request):
    categories               =   Category.objects.filter(is_blocked=True)
    if request.method == 'POST':
        category_id          =     request.POST.get('category')
        discount_percentage  =     request.POST['discount_percentage']
        category             =     Category.objects.get(id=category_id) 
        if discount_percentage.strip() == '':
            messages.error(request, "Discount percentage name cannot be null")
            return redirect('add_category_offers')
        elif Offer.objects.filter(category=category).exists():
            messages.error(request,"Category '{}' already have offers...".format(category.category_name))
            return redirect('add_category_offers')

        try:
            offer=Offer(discount_percentage=discount_percentage,category=category) 
            offer.save()
            messages.success(request,'Offer added successfully..')
            return redirect('add_category_offers')
        except IntegrityError:
            error_message =  'Category with name "{}" already exists'.format(category.category_name)
            messages.error(request, error_message)
            return render(request, 'admin/add_category_offers.html', {'error': error_message})
    else:
        context   =   {
                           'categories'   :   categories
                      }
        return render(request, 'admin/add_category_offers.html',context)
    
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)    
def display_offers(request): 
    offers      =  Offer.objects.all()
    context     =  {
        'offers'     :   offers,
    }
    return render(request,'admin/display_offers.html',context) 

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def block_category_offer(request):
    if request.method == 'POST':
        data          =   json.loads(request.body)
        offer_id      =   data.get('offerId')
        offer         =   Offer.objects.get(id=offer_id)
        if offer.is_deleted:
            offer.is_deleted    =   False
            offer.save()
        else:
            offer.is_deleted    =   True
            offer.save()
        response_data = {'message': 'Category offer blocked successfully','is_deleted':offer.is_deleted}
        return JsonResponse(response_data, status=200)
    else:
        response_data = {'error': 'Invalid request method'}
        return JsonResponse(response_data, status=400)
    
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def get_products(request):
    logging.debug('get_products function called')
    category_id = request.GET.get('category_id')
    products = Product.objects.filter(category_id=category_id,is_deleted=False)
    data = []
    for product in products:
        data.append({
            'id': product.id,
            'product_name': product.product_name
        })
    return JsonResponse(data, safe=False)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
def add_product_offers(request):
    if request.method =='POST':
        category_id           =   request.POST.get('category')
        product_id            =   request.POST.get('product')
        discount_percentage   =   request.POST.get('discount_percentage')
        category              =   Category.objects.get(id=category_id)
        products              =   Product.objects.get(id=product_id)
        if discount_percentage.strip()=='':
            messages.error(request,'Offer percentage  cannot be null')
            return redirect('add_product_offers')
        elif Offer.objects.filter(product=products).exists():
            messages.error(request,'Product with this offer already exists')
            return redirect('add_product_offers')
        try:
            offer  =   Offer(discount_percentage=discount_percentage,product=products) 
            offer.save()
            messages.success(request,'Offer added successfully..')
            return redirect('add_product_offers')
        except IntegrityError:
            messages.error(request,'Product with this offer already exists')
            return redirect('add_product_offers')
    else:
        categories   =   Category.objects.filter(is_blocked=True)
        context      =   {
                             'categories'   :   categories,
                         }
        return render(request, 'admin/add_product_offers.html',context)

   
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)    
def sales_report(request):
    orders = Orders.objects.filter(is_ordered = True).order_by('-id')
    context = {
                   'orders'   :   orders,
              }
    return render(request,'admin/sales_report.html',context)

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def monthly_report(request,date):
    frmdate = date
    fm = [2023, frmdate, 1]
    todt = [2023,frmdate,28]
    month_name = datetime.datetime.strptime(str(frmdate), "%m").strftime("%B")
    orders = Orders.objects.filter(date_added__gte = datetime.date(fm[0],fm[1],fm[2]),date_added__lte=datetime.date(todt[0],todt[1],todt[2]),is_ordered =True)
    total = sum([order.total for order in orders])
    if len(orders)>0:
        context = {
            'frmdate'   : frmdate,
            'orders'    : orders,
            'month_name': month_name,
            'total'     : total,
        }
        return render(request,'admin/monthly_sales_report.html',context)

    else:
        messages.error(request,"Sorry no oders in this month......")
        return render(request,'admin/monthly_sales_report.html')

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def yearly_report(request,date):
    frmdate         =    date
    fm              =    [frmdate, 1, 1]
    todt            =    [frmdate,12,31]
    orders   =    Orders.objects.filter(date_added__gte = datetime.date(fm[0],fm[1],fm[2]),date_added__lte=datetime.date(todt[0],todt[1],todt[2]),is_ordered =True)
    total = sum([order.total for order in orders])
    if len(orders)>0:
        context     =    {
                                'orders'  :  orders,
                                'year'           :  frmdate,
                                'total'     : total,
                         }
        return render(request,'admin/yearly_sales_report.html',context)
    else:
        messages.error(request,"Sorry no orders in this year...")
        return render(request,'admin/yearly_sales_report.html') 
      
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def date_range(request):
    if request.method == "POST":
        from_date = request.POST.get('from-date')
        to_date = request.POST.get('to-date')
        if from_date and to_date:
            date_from = from_date.split("-")
            date_to = to_date.split("-")
            fm = [int(x) for x in date_from]
            todt = [int(x) for x in date_to]
            orders = Orders.objects.filter(date_added__gte=datetime.date(fm[0], fm[1], fm[2]), date_added__lte=datetime.date(todt[0], todt[1], todt[2]), is_ordered=True)
            total = sum([order.total for order in orders])
            if orders:
                context = {
                    'orders': orders,
                    'from_date': from_date,
                    'to_date': to_date,
                    'total'     : total,
                }
                return render(request, 'admin/date_range_report.html', context)
            else:
                messages.error(request, "Sorry, no orders found for the selected date range.")
        else:
            messages.error(request, "Please select a valid date range.")
    return render(request, 'admin/sales_report.html')


@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def monthly_sales_pdf_download(request, frmdate):
    fm = [2023, int(frmdate), 1]
    todt = [2023, int(frmdate), 28]
    orders = Orders.objects.filter(date_added__gte=datetime.date(fm[0], fm[1], fm[2]), date_added__lte=datetime.date(todt[0], todt[1], todt[2]), is_ordered=True)
    total = sum([order.total for order in orders])
    if len(orders) > 0:
        template = get_template('admin/pdf_monthly_sales_report.html')
        context = {
            'frmdate': frmdate,
            'orders': orders,
            'total'     : total,
        }
        html = template.render(context)

        def render_to_pdf(html_string):
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
            if not pdf.err:
                return result.getvalue()
            return None

        pdf_file = render_to_pdf(html)
        if pdf_file:
            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"monthly_sales_report_{frmdate}.pdf"
            content = f"attachment; filename={filename}"
            response['Content-Disposition'] = content
            return response
    else:
        messages.error(request, "Sorry, no orders in this month.")
    return render(request, 'admin/monthly_sales_report.html')

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def yearly_sales_pdf_download(request,year):
    frmdate         =    year
    fm              =    [frmdate, 1, 1]
    todt            =    [frmdate,12,31]
    orders   =    Orders.objects.filter(date_added__gte = datetime.date(fm[0],fm[1],fm[2]),date_added__lte=datetime.date(todt[0],todt[1],todt[2]),is_ordered =True)
    total = sum([order.total for order in orders])
    if len(orders)>0:
        template = get_template('admin/pdf_yearly_sales_report.html')
        context     =    {
                                'orders'         :  orders,
                                'year'           :  frmdate,
                                'total'     : total,
                         }
        html = template.render(context)
        def render_to_pdf(html_string):
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
            if not pdf.err:
                return result.getvalue()
            return None

        pdf_file = render_to_pdf(html)
        if pdf_file:
            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"yearly_sales_report_{frmdate}.pdf"
            content = f"attachment; filename={filename}"
            response['Content-Disposition'] = content
            return response
    else:
        messages.error(request, "Sorry, no orders in this year.")
    return render(request, 'admin/yearly_sales_report.html')      

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def date_range_sales_pdf_download(request,from_date,to_date):
    fm = [int(x) for x in from_date.split("-")]
    todt = [int(x) for x in to_date.split("-")]

    orders = Orders.objects.filter(date_added__gte=datetime.date(fm[0], fm[1], fm[2]),
                                   date_added__lte=datetime.date(todt[0], todt[1], todt[2]), is_ordered=True)
    total = sum([order.total for order in orders])
    if len(orders) > 0:
        template = get_template('admin/pdf_date_range_report.html')
        context = {
            'orders': orders,
            'from_date': from_date,
            'to_date': to_date,
            'total'     : total,
        }
        html = template.render(context)

        def render_to_pdf(html_string):
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
            if not pdf.err:
                return result.getvalue()
            return None

        pdf_file = render_to_pdf(html)
        if pdf_file:
            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"Sales_report_{from_date}_{to_date}.pdf"
            content = f"attachment; filename={filename}"
            response['Content-Disposition'] = content
            return response
    else:
        messages.error(request, "Sorry, no orders in this date range.")
        return redirect('admin:sales_report')    
    
@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def monthly_sales_excel_download(request, frmdate):
    fm = [2023, int(frmdate), 1]
    todt = [2023, int(frmdate), 28]
    orders = Orders.objects.filter(date_added__gte=datetime.date(fm[0], fm[1], fm[2]), date_added__lte=datetime.date(todt[0], todt[1], todt[2]), is_ordered=True)
    total = sum([order.total for order in orders])
    if len(orders) > 0:
        # Create a new workbook and set the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        fm_str = '-'.join(str(x) for x in fm)
        worksheet.title = f"Monthly Sales Report ({fm_str})"
        # Set column widths and alignment
        column_widths = [25, 32, 32, 32, 35, 30,30,30,35]   
        alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')     
        for i, column_width in enumerate(column_widths):
            column_letter = get_column_letter(i+1)
            worksheet.column_dimensions[column_letter].width = column_width
            for cell in worksheet[column_letter]:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Add header row
        header_row = ['Date','Order Number', 'Customer Name','Phone Number','Address','Payment Method', 'Payment Status', 'Order Status', 'Total  Amount']
        for i, header_text in enumerate(header_row):
            cell = worksheet.cell(row=1, column=i+1)
            cell.value = header_text
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Add data rows
        for i, order in enumerate(orders):
            row_index = i+2
            worksheet.cell(row=row_index, column=1).value = order.date_added
            worksheet.cell(row=row_index, column=2).value = order.orderid
            worksheet.cell(row=row_index, column=3).value = order.address.firstname + ' ' + order.address.lastname
            worksheet.cell(row=row_index, column=4).value = order.address.phonenumber
            # address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            # worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)  
            address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)
    # Set cell alignment to center
            for col in range(1, 9):
                cell = worksheet.cell(row=row_index, column=col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # Set row height for address cells
            address_rows = worksheet[row_index]
            for cell in address_rows:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
                cell.font = openpyxl.styles.Font(size=10)
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            worksheet.row_dimensions[row_index].height = 45           
            worksheet.cell(row=row_index, column=6).value = order.payment.payment_method 
            worksheet.cell(row=row_index, column=7).value = order.payment.status 
            worksheet.cell(row=row_index, column=8).value = order.status
            worksheet.cell(row=row_index, column=9).value = order.total 
            last_row = worksheet.max_row
            worksheet.cell(row=last_row+1, column=8).value = "Total"
            worksheet.cell(row=last_row+1, column=9).value = total
            worksheet.cell(row=last_row+1, column=9).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=last_row+1, column=8).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
            for cell in worksheet[row_index]:
                cell.alignment = alignment
        # Save workbook to BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Set response headers
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="monthly_sales_report_{frmdate}.xlsx"'
        
        return response
    return HttpResponse("No data available for the selected month")

@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def yearly_sales_excel_download(request, year):
    fm              =    [year, 1, 1]
    todt            =    [year,12,31]
    orders   =    Orders.objects.filter(date_added__gte = datetime.date(fm[0],fm[1],fm[2]),date_added__lte=datetime.date(todt[0],todt[1],todt[2]),is_ordered =True)
    total = sum([order.total for order in orders])
    if len(orders) > 0:
        # Create a new workbook and set the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        fm_str = '-'.join(str(x) for x in fm)
        worksheet.title = f"Yearly Sales Report ({year})"
        # Set column widths and alignment
        column_widths = [25, 32, 32, 32, 35, 30,30,30,35]   
        alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')     
        for i, column_width in enumerate(column_widths):
            column_letter = get_column_letter(i+1)
            worksheet.column_dimensions[column_letter].width = column_width
            for cell in worksheet[column_letter]:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Add header row
        header_row = ['Date','Order Number', 'Customer Name','Phone Number','Address','Payment Method', 'Payment Status', 'Order Status', 'Total  Amount']
        for i, header_text in enumerate(header_row):
            cell = worksheet.cell(row=1, column=i+1)
            cell.value = header_text
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Add data rows
        for i, order in enumerate(orders):
            row_index = i+2
            worksheet.cell(row=row_index, column=1).value = order.date_added
            worksheet.cell(row=row_index, column=2).value = order.orderid
            worksheet.cell(row=row_index, column=3).value = order.address.firstname + ' ' + order.address.lastname
            worksheet.cell(row=row_index, column=4).value = order.address.phonenumber
            # address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            # worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)  
            address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)
    # Set cell alignment to center
            for col in range(1, 9):
                cell = worksheet.cell(row=row_index, column=col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # Set row height for address cells
            address_rows = worksheet[row_index]
            for cell in address_rows:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
                cell.font = openpyxl.styles.Font(size=10)
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            worksheet.row_dimensions[row_index].height = 45           
            worksheet.cell(row=row_index, column=6).value = order.payment.payment_method 
            worksheet.cell(row=row_index, column=7).value = order.payment.status 
            worksheet.cell(row=row_index, column=8).value = order.status
            worksheet.cell(row=row_index, column=9).value = order.total 
            last_row = worksheet.max_row
            worksheet.cell(row=last_row+1, column=8).value = "Total"
            worksheet.cell(row=last_row+1, column=9).value = total
            worksheet.cell(row=last_row+1, column=9).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=last_row+1, column=8).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
            for cell in worksheet[row_index]:
                cell.alignment = alignment
        # Save workbook to BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Set response headers
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Yearly_sales_report_{year}.xlsx"'
        
        return response
        
    return HttpResponse("No data available for the selected month")


@login_required(login_url='adminLogin') 
@permission_required('is_superuser')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def date_range_sales_excel_download(request, from_date,to_date):
    fm = [int(x) for x in from_date.split("-")]
    todt = [int(x) for x in to_date.split("-")]

    orders = Orders.objects.filter(date_added__gte=datetime.date(fm[0], fm[1], fm[2]),
                                   date_added__lte=datetime.date(todt[0], todt[1], todt[2]), is_ordered=True)
    total = sum([order.total for order in orders])
    if len(orders) > 0:
        # Create a new workbook and set the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        fm_str = '-'.join(str(x) for x in fm)
        worksheet.title = f"Yearly Sales Report ({from_date}_{to_date})"
        # Set column widths and alignment
        column_widths = [25, 32, 32, 32, 35, 30,30,30,35]   
        alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')     
        for i, column_width in enumerate(column_widths):
            column_letter = get_column_letter(i+1)
            worksheet.column_dimensions[column_letter].width = column_width
            for cell in worksheet[column_letter]:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Add header row
        header_row = ['Date','Order Number', 'Customer Name','Phone Number','Address','Payment Method', 'Payment Status', 'Order Status', 'Total  Amount']
        for i, header_text in enumerate(header_row):
            cell = worksheet.cell(row=1, column=i+1)
            cell.value = header_text
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Add data rows
        for i, order in enumerate(orders):
            row_index = i+2
            worksheet.cell(row=row_index, column=1).value = order.date_added
            worksheet.cell(row=row_index, column=2).value = order.orderid
            worksheet.cell(row=row_index, column=3).value = order.address.firstname + ' ' + order.address.lastname
            worksheet.cell(row=row_index, column=4).value = order.address.phonenumber
            # address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            # worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)  
            address_lines = [order.address.housename, order.address.locality, f"{order.address.city}, {order.address.state} {order.address.pincode}"]
            worksheet.cell(row=row_index, column=5).value =  "\n".join(address_lines)
    # Set cell alignment to center
            for col in range(1, 9):
                cell = worksheet.cell(row=row_index, column=col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # Set row height for address cells
            address_rows = worksheet[row_index]
            for cell in address_rows:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
                cell.font = openpyxl.styles.Font(size=10)
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            worksheet.row_dimensions[row_index].height = 45           
            worksheet.cell(row=row_index, column=6).value = order.payment.payment_method 
            worksheet.cell(row=row_index, column=7).value = order.payment.status 
            worksheet.cell(row=row_index, column=8).value = order.status
            worksheet.cell(row=row_index, column=9).value = order.total 
            last_row = worksheet.max_row
            worksheet.cell(row=last_row+1, column=8).value = "Total"
            worksheet.cell(row=last_row+1, column=9).value = total
            worksheet.cell(row=last_row+1, column=9).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=last_row+1, column=8).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
          
            for cell in worksheet[row_index]:
                cell.alignment = alignment
        # Save workbook to BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Set response headers
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Sales_report_{from_date}_{to_date}.xlsx"'
        
        return response
        
    return HttpResponse("No data available for the range...")